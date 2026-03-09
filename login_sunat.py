import json
import os
import re
import threading
import time
import tkinter as tk
from datetime import datetime, timedelta
from pathlib import Path
from tkinter import ttk, messagebox, font as tkfont
from urllib.parse import unquote, urlparse

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference
from pypdf import PdfReader
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
from tkcalendar import DateEntry


URL_LOGIN = (
    "https://api-seguridad.sunat.gob.pe/v1/clientessol/4f3b88b3-d9d6-402a-b85d-6a0bc857746a/"
    "oauth2/loginMenuSol?lang=es-PE&showDni=true&showLanguages=false&"
    "originalUrl=https://e-menu.sunat.gob.pe/cl-ti-itmenu/AutenticaMenuInternet.htm&"
    "state=rO0ABXNyABFqYXZhLnV0aWwuSGFzaE1hcAUH2sHDFmDRAwACRgAKbG9hZEZhY3RvckkACXRocmVzaG9sZHhwP0AAAAAAAAx3CAAAABAAAAADdAADZXhlcHQABnBhcmFtc3QASyomKiYvY2wtdGktaXRtZW51L01lbnVJbnRlcm5ldC5odG0mYjY0ZDI2YThiNWFmMDkxOTIzYjIzYjY0MDdhMWMxZGI0MWU3MzNhNnQABGV4ZWNweA=="
)

RUTA_CREDENCIALES = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "credenciales.json"
)

PATRON_FECHA = re.compile(r"^\d{2}/\d{2}/\d{4}$")

MESES_ES = {
    1: "ENERO",
    2: "FEBRERO",
    3: "MARZO",
    4: "ABRIL",
    5: "MAYO",
    6: "JUNIO",
    7: "JULIO",
    8: "AGOSTO",
    9: "SETIEMBRE",
    10: "OCTUBRE",
    11: "NOVIEMBRE",
    12: "DICIEMBRE",
}

MESES_LISTA = [
    "ENERO",
    "FEBRERO",
    "MARZO",
    "ABRIL",
    "MAYO",
    "JUNIO",
    "JULIO",
    "AGOSTO",
    "SETIEMBRE",
    "OCTUBRE",
    "NOVIEMBRE",
    "DICIEMBRE",
]

MESES_IDX = {m: i + 1 for i, m in enumerate(MESES_LISTA)}


def _rango_mensual(desde_mes: int, desde_anio: int, hasta_mes: int, hasta_anio: int):
    """
    Genera tuplas (fecha_inicio_str, fecha_fin_str, anio, mes_nombre, mes_num)
    en orden cronológico, con meses completos.
    """
    start = datetime(desde_anio, desde_mes, 1)
    end = datetime(hasta_anio, hasta_mes, 1)
    if start > end:
        start, end = end, start

    cur = start
    while cur <= end:
        anio = cur.year
        mes_num = cur.month
        mes_nombre = MESES_ES.get(mes_num, str(mes_num))
        if mes_num == 12:
            next_month = datetime(anio + 1, 1, 1)
        else:
            next_month = datetime(anio, mes_num + 1, 1)
        last_day = next_month - timedelta(days=1)
        yield (
            cur.strftime("%d/%m/%Y"),
            last_day.strftime("%d/%m/%Y"),
            str(anio),
            mes_nombre,
            mes_num,
        )
        cur = next_month


def validar_fecha(texto):
    if not texto or not texto.strip():
        return False
    texto = texto.strip()
    if not PATRON_FECHA.match(texto):
        return False
    try:
        datetime.strptime(texto, "%d/%m/%Y")
        return True
    except ValueError:
        return False


def cargar_credenciales_desde_archivo():
    if not os.path.exists(RUTA_CREDENCIALES):
        return None
    try:
        with open(RUTA_CREDENCIALES, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None


def guardar_credenciales(ruc, usuario, contrasena):
    with open(RUTA_CREDENCIALES, "w", encoding="utf-8") as f:
        json.dump(
            {
                "ruc": ruc,
                "usuario": usuario,
                "contrasena": contrasena
            },
            f,
            indent=4,
            ensure_ascii=False
        )


def esta_visible(locator):
    try:
        return locator.count() > 0 and locator.first.is_visible()
    except Exception:
        return False


def click_primer_visible(page_or_frame, selectores, log=None, espera_ms=1200, force=False):
    for tipo, valor in selectores:
        try:
            if tipo == "locator":
                obj = page_or_frame.locator(valor)
            elif tipo == "text":
                obj = page_or_frame.get_by_text(valor, exact=True)
                if obj.count() == 0:
                    obj = page_or_frame.get_by_text(valor)
            elif tipo == "role_button":
                obj = page_or_frame.get_by_role("button", name=valor)
            else:
                continue

            if obj.count() > 0:
                try:
                    obj.first.scroll_into_view_if_needed(timeout=2000)
                except Exception:
                    pass

                if force:
                    obj.first.click(force=True)
                else:
                    if obj.first.is_visible():
                        obj.first.click()
                    else:
                        continue

                page_or_frame.wait_for_timeout(espera_ms)
                if log:
                    log(f"Click OK: {valor}")
                return True
        except Exception:
            pass
    return False


def obtener_target_con_fechas(page):
    selector_desde = '[id="criterio.fec_desde"]'
    selector_hasta = '[id="criterio.fec_hasta"]'

    try:
        if page.locator(selector_desde).count() > 0 and page.locator(selector_hasta).count() > 0:
            return page
    except Exception:
        pass

    for frame in page.frames:
        try:
            if frame.locator(selector_desde).count() > 0 and frame.locator(selector_hasta).count() > 0:
                return frame
        except Exception:
            pass

    return None


def localizar_contenedor_resultados(page):
    candidatos = [page] + list(page.frames)

    for obj in candidatos:
        try:
            if obj.locator("a:has-text('Descargar PDF')").count() > 0:
                return obj
        except Exception:
            pass

    for obj in candidatos:
        try:
            if obj.get_by_text("Facturas Electrónicas Emitidas del Periodo").count() > 0:
                return obj
        except Exception:
            pass

    return None


def esperar_resultados_consulta(page, log=None, timeout_seg=60):
    if log:
        log("Esperando resultados reales de la consulta...")

    fin = time.time() + timeout_seg

    while time.time() < fin:
        cont = localizar_contenedor_resultados(page)
        if cont is not None:
            try:
                if cont.locator("a:has-text('Descargar PDF')").count() > 0:
                    if log:
                        log("Resultados detectados: se encontraron enlaces 'Descargar PDF'.")
                    return cont
            except Exception:
                pass

            try:
                if cont.get_by_text("Facturas Electrónicas Emitidas del Periodo").count() > 0:
                    if log:
                        log("Pantalla de resultados detectada.")
                    return cont
            except Exception:
                pass

        time.sleep(0.8)

    raise Exception(
        "La consulta sí se envió, pero no se detectaron a tiempo los resultados ni los enlaces 'Descargar PDF'."
    )




def mantener_navegador_abierto(browser, log_callback=None):
    def log(msg):
        if log_callback:
            log_callback(msg)
        else:
            print(msg)

    log("El navegador quedará abierto hasta que lo cierres manualmente.")

    while True:
        try:
            if not browser.is_connected():
                break
            time.sleep(1)
        except Exception:
            break

    log("Navegador cerrado por el usuario.")


def _sanear_nombre_archivo(nombre: str, max_len: int = 140) -> str:
    nombre = (nombre or "").strip()
    nombre = re.sub(r'[<>:"/\\|?*\n\r\t]', " ", nombre)
    nombre = re.sub(r"\s+", " ", nombre).strip()
    if len(nombre) > max_len:
        nombre = nombre[:max_len].rstrip()
    return nombre


def _leer_texto_pdf(ruta_pdf: Path) -> str:
    lector = PdfReader(str(ruta_pdf))
    partes = []
    for page_pdf in lector.pages:
        try:
            partes.append(page_pdf.extract_text() or "")
        except Exception:
            partes.append("")
    return "\n".join(partes)


def extraer_datos_factura_pdf(ruta_pdf: Path) -> dict:
    texto = _leer_texto_pdf(ruta_pdf)

    m_serie = re.search(r"\b([A-Z]\d{3})-(\d+)\b", texto)
    serie = m_serie.group(1) if m_serie else ""
    numero = m_serie.group(2) if m_serie else ""

    m_fecha = re.search(r"Fecha de Emisi[oó]n\s*:\s*(\d{2}/\d{2}/\d{4})", texto, re.IGNORECASE)
    fecha_emision = m_fecha.group(1) if m_fecha else ""

    m_emp = re.search(r"Señor\(es\)\s*:\s*(.+)", texto)
    empresa = m_emp.group(1).strip() if m_emp else ""

    m_mon = re.search(r"Tipo de Moneda\s*:\s*(.+)", texto, re.IGNORECASE)
    moneda = m_mon.group(1).strip().upper() if m_mon else ""

    items = []
    lineas = [ln.strip() for ln in texto.splitlines() if ln.strip()]
    idx = None
    for i, ln in enumerate(lineas):
        if ln.startswith("Cantidad") and "Descripción" in ln and "Valor Unitario" in ln:
            idx = i + 1
            break

    if idx is not None:
        # Collect all item lines between header and footer
        item_lines = []
        _STOP_PREFIXES = (
            "Valor de Venta", "SON:", "Sub Total", "ISC", "IGV",
            "ICBPER", "Importe Total", "Anticipos", "Descuentos",
            "Otros Cargos", "Otros Tributos", "Monto de redondeo",
        )
        for ln in lineas[idx:]:
            if any(ln.startswith(p) for p in _STOP_PREFIXES):
                break
            item_lines.append(ln)

        # Parse items: each item starts with a line like "30.00 UNIDAD description... [price]"
        # Description may continue on next lines, and price may be on its own line
        pending_cantidad = None
        pending_unidad = ""
        pending_desc_parts = []

        def _try_float(s: str):
            try:
                return float(s.replace(",", ""))
            except Exception:
                return None

        def _flush_item(cantidad, desc_parts, p_unit):
            desc = " ".join(desc_parts).strip()
            total = round(cantidad * p_unit, 2)
            items.append({
                "cantidad": cantidad,
                "descripcion": desc,
                "p_unit": p_unit,
                "total": total,
            })

        for ln in item_lines:
            partes = ln.split()
            if not partes:
                continue

            # Check if this line starts a new item (number + UNIDAD/unit)
            first_val = _try_float(partes[0])
            is_new_item = False
            if first_val is not None and len(partes) >= 2:
                # Check if second word looks like a unit of measure
                second = partes[1].upper()
                known_units = (
                    "UNIDAD", "UND", "UNI", "UNID", "KG", "KGS",
                    "LT", "LTS", "LITRO", "LITROS", "MT", "MTS",
                    "METRO", "METROS", "CAJA", "CAJAS", "ROLLO",
                    "ROLLOS", "PAR", "PARES", "JUEGO", "PIEZA",
                    "PIEZAS", "BIDON", "BIDONES", "BALDE", "GALÓN",
                    "GALON", "BOLSA", "BOLSAS", "PAQUETE", "MILLAR",
                    "CIENTO", "DOCENA", "TONELADA", "SERVICIO",
                    "GLOBAL", "GLB",
                )
                if second in known_units or len(second) <= 5:
                    is_new_item = True

            if is_new_item:
                # If there was a pending item missing its price, try the last part of this line
                if pending_cantidad is not None and pending_desc_parts:
                    # Price was never found; skip the pending item or look harder
                    # Actually check if any trailing number on desc_parts is the price
                    last = pending_desc_parts[-1] if pending_desc_parts else ""
                    lv = _try_float(last)
                    if lv is not None and lv < pending_cantidad * 10000:
                        pending_desc_parts.pop()
                        _flush_item(pending_cantidad, pending_desc_parts, lv)
                    # else: item without price, skip

                pending_cantidad = first_val
                pending_unidad = partes[1] if len(partes) > 1 else ""
                rest = partes[2:]  # description + maybe price at end

                # Check if last part of this line is the price
                if rest:
                    last_val = _try_float(rest[-1])
                    if last_val is not None and len(rest) > 1:
                        # Price is on same line
                        desc = " ".join(rest[:-1])
                        _flush_item(pending_cantidad, [desc], last_val)
                        pending_cantidad = None
                        pending_desc_parts = []
                        continue

                pending_desc_parts = list(rest)

            elif pending_cantidad is not None:
                # Continuation line: could be more description or a standalone price
                if len(partes) == 1:
                    val = _try_float(partes[0])
                    if val is not None:
                        # This is the price on its own line
                        _flush_item(pending_cantidad, pending_desc_parts, val)
                        pending_cantidad = None
                        pending_desc_parts = []
                        continue

                # Check if last token is the price
                last_val = _try_float(partes[-1])
                if last_val is not None and len(partes) > 1:
                    pending_desc_parts.extend(partes[:-1])
                    _flush_item(pending_cantidad, pending_desc_parts, last_val)
                    pending_cantidad = None
                    pending_desc_parts = []
                else:
                    # Pure description continuation
                    pending_desc_parts.extend(partes)

        # Flush any remaining pending item
        if pending_cantidad is not None and pending_desc_parts:
            last = pending_desc_parts[-1]
            lv = _try_float(last)
            if lv is not None:
                pending_desc_parts.pop()
                _flush_item(pending_cantidad, pending_desc_parts, lv)

    return {
        "fecha_emision": fecha_emision,
        "empresa": empresa,
        "moneda": moneda,
        "serie": serie,
        "numero": numero,
        "items": items,
    }


def _aplicar_headers(ws, headers):
    ws.append(headers)
    bold = Font(bold=True)
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = bold


def _ajustar_columnas_extraccion(ws):
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 60
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 70


def guardar_excel_por_anio_y_analisis(ruta_excel: Path, filas: list[dict]) -> None:
    """
    - 1 hoja por año (por ejemplo: '2025', '2026')
    - hoja 'ANALISIS' con tablas agregadas + gráficos
    """
    wb = Workbook()
    wb.remove(wb.active)

    headers = ["FECHA", "EMPRESA", "CANTIDAD", "DESCRIPCION", "MONEDA", "P_UNIT", "TOTAL", "URL"]

    filas_por_anio: dict[str, list[dict]] = {}
    for f in filas:
        fecha = (f.get("fecha") or "").strip()
        anio = ""
        if fecha and validar_fecha(fecha):
            try:
                anio = str(datetime.strptime(fecha, "%d/%m/%Y").year)
            except Exception:
                anio = ""
        if not anio:
            try:
                p = Path(f.get("url", ""))
                parts = list(p.parts)
                for j in range(len(parts) - 1):
                    if parts[j].upper() == "FACTURAS" and re.match(r"^\d{4}$", parts[j + 1]):
                        anio = parts[j + 1]
                        break
            except Exception:
                pass
        anio = anio or "SIN_ANIO"
        filas_por_anio.setdefault(anio, []).append(f)

    # hojas por año
    for anio in sorted(filas_por_anio.keys()):
        ws = wb.create_sheet(title=anio)
        _aplicar_headers(ws, headers)
        for f in filas_por_anio[anio]:
            ws.append(
                [
                    f.get("fecha", ""),
                    f.get("empresa", ""),
                    f.get("cantidad", ""),
                    f.get("descripcion", ""),
                    f.get("moneda", ""),
                    f.get("p_unit", ""),
                    f.get("total", ""),
                    f.get("url", ""),
                ]
            )
            r = ws.max_row
            cell = ws.cell(row=r, column=8)
            ruta = f.get("url", "")
            if ruta:
                try:
                    cell.hyperlink = Path(ruta).resolve().as_uri()
                except Exception:
                    cell.hyperlink = ruta
                cell.font = Font(color="0000FF", underline="single")
        _ajustar_columnas_extraccion(ws)

    # ANALISIS
    ws_a = wb.create_sheet(title="ANALISIS")
    ws_a["A1"] = "ANALISIS (resumen tipo tabla dinámica)"
    ws_a["A1"].font = Font(bold=True, size=12)

    # --- Tabla dinámica completa: AÑO, MES, CLIENTE, PRODUCTO, TOTAL ---
    pivot_header_row = 3
    pivot_headers = ["AÑO", "MES", "CLIENTE", "PRODUCTO", "TOTAL"]
    for ci, h in enumerate(pivot_headers, start=1):
        cell = ws_a.cell(row=pivot_header_row, column=ci, value=h)
        cell.font = Font(bold=True)

    # Agregar datos: (año, mes, cliente, producto) -> total
    pivot_data: dict[tuple[int, int, str, str], float] = {}
    for f in filas:
        fecha = (f.get("fecha") or "").strip()
        if not (fecha and validar_fecha(fecha)):
            continue
        dt = datetime.strptime(fecha, "%d/%m/%Y")
        y = dt.year
        m = dt.month
        cliente = (f.get("empresa") or "").strip()
        prod = (f.get("descripcion") or "").strip()
        try:
            total = float(f.get("total") or 0)
        except Exception:
            total = 0.0
        key = (y, m, cliente, prod)
        pivot_data[key] = pivot_data.get(key, 0.0) + total

    pivot_row = pivot_header_row + 1
    for (y, m, cli, prod) in sorted(pivot_data.keys()):
        ws_a.cell(row=pivot_row, column=1, value=y)
        ws_a.cell(row=pivot_row, column=2, value=MESES_ES.get(m, str(m)))
        ws_a.cell(row=pivot_row, column=3, value=cli)
        ws_a.cell(row=pivot_row, column=4, value=prod)
        ws_a.cell(row=pivot_row, column=5, value=round(pivot_data[(y, m, cli, prod)], 2))
        pivot_row += 1

    pivot_end_row = pivot_row - 1

    # --- TOP Cliente por período ---
    top_cli_start = pivot_row + 2
    ws_a.cell(row=top_cli_start, column=1, value="¿Qué cliente compra más por mes y año?").font = Font(bold=True, size=11)
    top_cli_header = top_cli_start + 1
    for ci, h in enumerate(["AÑO", "MES", "TOP_CLIENTE", "TOTAL"], start=1):
        ws_a.cell(row=top_cli_header, column=ci, value=h).font = Font(bold=True)

    tot_cliente: dict[tuple[int, int, str], float] = {}
    for (y, m, cli, _prod), tot in pivot_data.items():
        k = (y, m, cli)
        tot_cliente[k] = tot_cliente.get(k, 0.0) + tot

    top_cliente: dict[tuple[int, int], tuple[str, float]] = {}
    for (y, m, cli), tot in tot_cliente.items():
        key = (y, m)
        cur = top_cliente.get(key)
        if cur is None or tot > cur[1]:
            top_cliente[key] = (cli, tot)

    cli_data_start = top_cli_header + 1
    r = cli_data_start
    for (y, m) in sorted(top_cliente.keys()):
        cli, tot = top_cliente[(y, m)]
        ws_a.cell(row=r, column=1, value=y)
        ws_a.cell(row=r, column=2, value=MESES_ES.get(m, str(m)))
        ws_a.cell(row=r, column=3, value=cli)
        ws_a.cell(row=r, column=4, value=round(tot, 2))
        r += 1
    cli_data_end = r - 1

    # --- TOP Descripción por período ---
    top_prod_start = r + 2
    ws_a.cell(row=top_prod_start, column=1, value="¿Cuál descripción se vende más por mes y año?").font = Font(bold=True, size=11)
    top_prod_header = top_prod_start + 1
    for ci, h in enumerate(["AÑO", "MES", "TOP_DESCRIPCION", "TOTAL"], start=1):
        ws_a.cell(row=top_prod_header, column=ci, value=h).font = Font(bold=True)

    tot_prod: dict[tuple[int, int, str], float] = {}
    for (y, m, _cli, prod), tot in pivot_data.items():
        k = (y, m, prod)
        tot_prod[k] = tot_prod.get(k, 0.0) + tot

    top_prod: dict[tuple[int, int], tuple[str, float]] = {}
    for (y, m, prod), tot in tot_prod.items():
        key = (y, m)
        cur = top_prod.get(key)
        if cur is None or tot > cur[1]:
            top_prod[key] = (prod, tot)

    prod_data_start = top_prod_header + 1
    rp = prod_data_start
    for (y, m) in sorted(top_prod.keys()):
        prod, tot = top_prod[(y, m)]
        ws_a.cell(row=rp, column=1, value=y)
        ws_a.cell(row=rp, column=2, value=MESES_ES.get(m, str(m)))
        ws_a.cell(row=rp, column=3, value=prod)
        ws_a.cell(row=rp, column=4, value=round(tot, 2))
        rp += 1
    prod_data_end = rp - 1

    # --- Gráficos ---
    chart_row = rp + 2

    try:
        chart1 = BarChart()
        chart1.title = "¿Qué cliente compra más por mes y año?"
        chart1.y_axis.title = "Total"
        chart1.x_axis.title = "Período"
        chart1.width = 28
        chart1.height = 14
        data1 = Reference(ws_a, min_col=4, min_row=top_cli_header, max_row=cli_data_end)
        cats1 = Reference(ws_a, min_col=2, min_row=cli_data_start, max_row=cli_data_end)
        chart1.add_data(data1, titles_from_data=True)
        chart1.set_categories(cats1)
        ws_a.add_chart(chart1, f"A{chart_row}")
    except Exception:
        pass

    try:
        chart2 = BarChart()
        chart2.title = "¿Cuál descripción se vende más por mes y año?"
        chart2.y_axis.title = "Total"
        chart2.x_axis.title = "Período"
        chart2.width = 28
        chart2.height = 14
        data2 = Reference(ws_a, min_col=4, min_row=top_prod_header, max_row=prod_data_end)
        cats2 = Reference(ws_a, min_col=2, min_row=prod_data_start, max_row=prod_data_end)
        chart2.add_data(data2, titles_from_data=True)
        chart2.set_categories(cats2)
        ws_a.add_chart(chart2, f"A{chart_row + 16}")
    except Exception:
        pass

    ws_a.column_dimensions["A"].width = 10
    ws_a.column_dimensions["B"].width = 14
    ws_a.column_dimensions["C"].width = 40
    ws_a.column_dimensions["D"].width = 50
    ws_a.column_dimensions["E"].width = 14

    ruta_excel.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(ruta_excel))




def obtener_links_descargar_pdf(contenedor):
    try:
        links = contenedor.locator("a:has-text('Descargar PDF')")
        if links.count() > 0:
            return links
    except Exception:
        pass

    try:
        links = contenedor.get_by_text("Descargar PDF", exact=True)
        if links.count() > 0:
            return links
    except Exception:
        pass

    return None


def _es_respuesta_pdf(response):
    try:
        headers = response.headers
        ct = headers.get("content-type", "").lower()
        cd = headers.get("content-disposition", "").lower()
        url = response.url.lower()

        return (
            "application/pdf" in ct
            or "application/octet-stream" in ct
            or ".pdf" in url
            or ".pdf" in cd
        )
    except Exception:
        return False


def _nombre_desde_headers_o_url(headers: dict, url: str, default_name: str) -> str:
    cd = headers.get("content-disposition", "") or headers.get("Content-Disposition", "")
    if cd:
        m = re.search(r'filename\*?=(?:UTF-8\'\')?"?([^";]+)"?', cd, re.IGNORECASE)
        if m:
            nombre = unquote(m.group(1)).strip()
            nombre = nombre.replace('"', "").strip()
            if nombre.lower().endswith(".pdf"):
                return _sanear_nombre_archivo(nombre)

    try:
        parsed = urlparse(url)
        base = Path(parsed.path).name
        if base and base.lower().endswith(".pdf"):
            return _sanear_nombre_archivo(unquote(base))
    except Exception:
        pass

    return default_name


def _safe_click_link(locator):
    try:
        locator.scroll_into_view_if_needed(timeout=3000)
    except Exception:
        pass

    try:
        locator.click(force=True, timeout=5000)
        return
    except Exception:
        pass

    try:
        locator.evaluate("(el) => el.click()")
        return
    except Exception:
        pass

    try:
        locator.dispatch_event("click")
        return
    except Exception:
        pass

    raise Exception("No se pudo hacer clic en el enlace del PDF.")


def descargar_pdf_desde_link(page, contenedor_resultados, idx, tmp_dir: Path, log):
    def reacquire():
        links = obtener_links_descargar_pdf(contenedor_resultados)
        if links is None or links.count() <= idx:
            raise Exception(f"No se pudo reubicar el link PDF índice {idx}.")
        return links.nth(idx)

    nombre_base = f"factura_{idx + 1}.pdf"

    # ESTRATEGIA 1: download nativo
    try:
        link = reacquire()
        with page.expect_download(timeout=20000) as d:
            _safe_click_link(link)
        download = d.value
        sugerido = download.suggested_filename or nombre_base
        ruta_tmp = tmp_dir / _sanear_nombre_archivo(sugerido)
        download.save_as(str(ruta_tmp))
        log(f"PDF {idx + 1}: descargado por evento download.")
        return ruta_tmp
    except PlaywrightTimeoutError:
        log(f"PDF {idx + 1}: no hubo evento download. Probando captura de respuesta PDF...")
    except Exception as e:
        log(f"PDF {idx + 1}: fallo estrategia download ({e}). Probando respuesta PDF...")

    time.sleep(1)

    # ESTRATEGIA 2: capturar respuesta PDF
    try:
        link = reacquire()
        with page.expect_response(lambda r: _es_respuesta_pdf(r), timeout=25000) as rinfo:
            _safe_click_link(link)

        resp = rinfo.value
        data = resp.body()
        nombre = _nombre_desde_headers_o_url(resp.headers, resp.url, nombre_base)
        ruta_tmp = tmp_dir / nombre
        ruta_tmp.write_bytes(data)
        log(f"PDF {idx + 1}: guardado desde respuesta HTTP PDF.")
        return ruta_tmp
    except PlaywrightTimeoutError:
        log(f"PDF {idx + 1}: no se detectó respuesta PDF. Probando popup...")
    except Exception as e:
        log(f"PDF {idx + 1}: fallo estrategia respuesta ({e}). Probando popup...")

    time.sleep(1)

    # ESTRATEGIA 3: popup / nueva pestaña
    try:
        link = reacquire()
        with page.expect_popup(timeout=20000) as pinfo:
            _safe_click_link(link)

        popup = pinfo.value
        popup.wait_for_timeout(4000)

        url_pdf = popup.url

        if not url_pdf or url_pdf.startswith("about:blank"):
            try:
                emb = popup.locator("embed").first
                if emb.count() > 0:
                    url_pdf = emb.get_attribute("src")
            except Exception:
                pass

        if not url_pdf:
            raise Exception("El popup no expuso una URL utilizable del PDF.")

        resp = page.request.get(url_pdf, timeout=60000)
        if not resp.ok:
            raise Exception(f"No se pudo descargar el PDF del popup. Status: {resp.status}")

        headers = resp.headers
        nombre = _nombre_desde_headers_o_url(headers, url_pdf, nombre_base)
        ruta_tmp = tmp_dir / nombre
        ruta_tmp.write_bytes(resp.body())

        try:
            popup.close()
        except Exception:
            pass

        log(f"PDF {idx + 1}: descargado desde popup.")
        return ruta_tmp

    except PlaywrightTimeoutError:
        raise Exception(
            f"No se pudo descargar el PDF {idx + 1}: SUNAT no lanzó download, ni respuesta PDF capturable, ni popup utilizable."
        )
    except Exception as e:
        raise Exception(f"No se pudo descargar el PDF {idx + 1}: {e}")


def login_sunat(page, ruc, usuario, contrasena, log):
    log("Abriendo página de login SUNAT...")
    page.goto(URL_LOGIN, wait_until="domcontentloaded", timeout=60000)
    page.wait_for_selector("#txtRuc", state="visible", timeout=20000)
    page.fill("#txtRuc", ruc)
    page.wait_for_timeout(800)

    locator_usuario = page.locator("#txtUsuario")
    locator_contrasena = page.locator("#txtContrasena")

    usuario_visible = esta_visible(locator_usuario)
    contrasena_visible = esta_visible(locator_contrasena)

    if usuario_visible and contrasena_visible:
        log("Formulario completo detectado. Se hará login directo.")
        locator_usuario.first.fill(usuario)
        locator_contrasena.first.fill(contrasena)

        ok_login = click_primer_visible(
            page,
            [
                ("role_button", "Iniciar sesión"),
                ("locator", "button:has-text('Iniciar sesión')"),
                ("locator", "input[type='submit']"),
                ("locator", "button[type='submit']"),
                ("text", "Iniciar sesión"),
            ],
            log=log,
            espera_ms=2500
        )
        if not ok_login:
            locator_contrasena.first.press("Enter")
            page.wait_for_timeout(2500)
        return

    log("No se detectó formulario completo. Se intentará paso intermedio.")
    ok_continuar = click_primer_visible(
        page,
        [
            ("role_button", "Continuar"),
            ("role_button", "Siguiente"),
            ("role_button", "Continuar con RUC"),
            ("locator", "button:has-text('Continuar')"),
            ("locator", "button:has-text('Siguiente')"),
            ("text", "Continuar"),
            ("text", "Siguiente"),
        ],
        log=log,
        espera_ms=1800
    )

    if not ok_continuar:
        raise Exception("No se encontró el botón de continuación ni el formulario completo de login.")

    page.wait_for_selector("#txtContrasena", state="visible", timeout=20000)

    if page.locator("#txtUsuario").count() > 0:
        try:
            if page.locator("#txtUsuario").first.is_visible():
                page.locator("#txtUsuario").first.fill(usuario)
        except Exception:
            pass

    page.locator("#txtContrasena").first.fill(contrasena)

    ok_login = click_primer_visible(
        page,
        [
            ("role_button", "Iniciar sesión"),
            ("locator", "button:has-text('Iniciar sesión')"),
            ("locator", "input[type='submit']"),
            ("locator", "button[type='submit']"),
            ("text", "Iniciar sesión"),
        ],
        log=log,
        espera_ms=2500
    )
    if not ok_login:
        page.locator("#txtContrasena").first.press("Enter")
        page.wait_for_timeout(2500)


def ejecutar_rpa(ruc, usuario, contrasena, fecha_inicio, fecha_fin, log_callback=None):
    def log(msg):
        if log_callback:
            log_callback(msg)
        else:
            print(msg)

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=False,
            slow_mo=250
        )

        base_facturas = Path(__file__).resolve().parent / "FACTURAS"
        base_facturas.mkdir(parents=True, exist_ok=True)

        context = browser.new_context(
            viewport={"width": 1366, "height": 900},
            locale="es-PE",
            accept_downloads=True,
        )

        page = context.new_page()

        try:
            login_sunat(page, ruc, usuario, contrasena, log)

            log("Esperando menú principal...")
            page.wait_for_timeout(3000)

            log("Ingresando a Empresas...")
            if not click_primer_visible(
                page,
                [
                    ("locator", "h4:has-text('Empresas')"),
                    ("text", "Empresas"),
                ],
                log=log
            ):
                raise Exception("No se encontró la opción 'Empresas'.")

            log("Ingresando a Comprobantes de pago...")
            if not click_primer_visible(
                page,
                [
                    ("locator", "span.spanNivelDescripcion:has-text('Comprobantes de pago')"),
                    ("text", "Comprobantes de pago"),
                ],
                log=log
            ):
                raise Exception("No se encontró la opción 'Comprobantes de pago'.")

            log("Ingresando a SEE - SOL...")
            if not click_primer_visible(
                page,
                [
                    ("locator", "span.spanNivelDescripcion:has-text('SEE - SOL')"),
                    ("text", "SEE - SOL"),
                ],
                log=log
            ):
                raise Exception("No se encontró la opción 'SEE - SOL'.")

            log("Ingresando a Factura Electrónica...")
            if not click_primer_visible(
                page,
                [
                    ("locator", "span.spanNivelDescripcion:has-text('Factura Electrónica')"),
                    ("text", "Factura Electrónica"),
                    ("locator", "li#nivel3_11_5_3"),
                ],
                log=log
            ):
                raise Exception("No se encontró la opción 'Factura Electrónica'.")

            log("Ingresando a Consultar Factura y Nota...")
            if not click_primer_visible(
                page,
                [
                    ("locator", "span.spanNivelDescripcion:has-text('Consultar Factura y Nota')"),
                    ("text", "Consultar Factura y Nota"),
                ],
                log=log,
                espera_ms=2500
            ):
                raise Exception("No se encontró la opción 'Consultar Factura y Nota'.")

            if len(context.pages) > 1:
                page = context.pages[-1]
                page.wait_for_timeout(2000)

            log("Buscando formulario de fechas...")
            target = obtener_target_con_fechas(page)
            if target is None:
                raise Exception("No se localizaron los campos de fecha.")

            log("Rellenando fechas...")
            target.locator('[id="criterio.fec_desde"]').first.fill(fecha_inicio)
            target.locator('[id="criterio.fec_hasta"]').first.fill(fecha_fin)
            target.wait_for_timeout(700)

            log("Haciendo clic en Aceptar...")
            aceptado = click_primer_visible(
                target,
                [
                    ("role_button", "Aceptar"),
                    ("locator", "button:has-text('Aceptar')"),
                    ("locator", ".dijitButton:has(.nextIcon)"),
                    ("locator", "button:has(span.nextIcon)"),
                    ("text", "Aceptar"),
                ],
                log=log,
                espera_ms=2500
            )

            if not aceptado and target != page:
                aceptado = click_primer_visible(
                    page,
                    [
                        ("role_button", "Aceptar"),
                        ("locator", "button:has-text('Aceptar')"),
                        ("locator", ".dijitButton:has(.nextIcon)"),
                        ("locator", "button:has(span.nextIcon)"),
                        ("text", "Aceptar"),
                    ],
                    log=log,
                    espera_ms=2500
                )

            if not aceptado:
                raise Exception("No se encontró el botón 'Aceptar'.")

            contenedor_resultados = esperar_resultados_consulta(page, log=log, timeout_seg=60)

            time.sleep(2)

            log("Buscando enlaces 'Descargar PDF'...")
            links = obtener_links_descargar_pdf(contenedor_resultados)
            if links is None or links.count() == 0:
                links = obtener_links_descargar_pdf(page)

            if links is None or links.count() == 0:
                raise Exception("No se encontraron enlaces 'Descargar PDF' en los resultados.")

            total_links = links.count()
            log(f"Se encontraron {total_links} PDF(s) para descargar.")

            filas_excel = []
            tmp_dir = base_facturas / "_tmp"
            tmp_dir.mkdir(parents=True, exist_ok=True)

            dt_inicio = datetime.strptime(fecha_inicio, "%d/%m/%Y")
            anio = str(dt_inicio.year)
            mes = MESES_ES.get(dt_inicio.month, str(dt_inicio.month))

            for i in range(total_links):
                log(f"Descargando PDF {i + 1}/{total_links}...")
                ruta_tmp = descargar_pdf_desde_link(page, contenedor_resultados, i, tmp_dir, log)

                datos = extraer_datos_factura_pdf(ruta_tmp)

                empresa = _sanear_nombre_archivo(datos.get("empresa", "SIN_EMPRESA"))
                serie = datos.get("serie", "")
                numero = datos.get("numero", "")
                fecha_emision = datos.get("fecha_emision", "")
                moneda = datos.get("moneda", "")

                carpeta_dest = base_facturas / anio / mes / empresa
                carpeta_dest.mkdir(parents=True, exist_ok=True)

                descripcion_primera = ""
                items = datos.get("items") or []
                if items:
                    descripcion_primera = items[0].get("descripcion", "")

                desc_safe = _sanear_nombre_archivo(descripcion_primera)
                if serie and numero and desc_safe:
                    nombre_final = f"{serie}_{numero}_{desc_safe}.pdf"
                elif serie and numero:
                    nombre_final = f"{serie}_{numero}.pdf"
                else:
                    nombre_final = _sanear_nombre_archivo(ruta_tmp.stem) + ".pdf"

                ruta_final = carpeta_dest / nombre_final

                if ruta_final.exists():
                    base = ruta_final.stem
                    ext = ruta_final.suffix
                    k = 2
                    while True:
                        cand = carpeta_dest / f"{base}_{k}{ext}"
                        if not cand.exists():
                            ruta_final = cand
                            break
                        k += 1

                ruta_tmp.replace(ruta_final)

                if items:
                    for it in items:
                        filas_excel.append(
                            {
                                "fecha": fecha_emision,
                                "empresa": datos.get("empresa", ""),
                                "cantidad": it.get("cantidad", ""),
                                "descripcion": it.get("descripcion", ""),
                                "moneda": moneda,
                                "p_unit": it.get("p_unit", ""),
                                "total": it.get("total", ""),
                                "url": str(ruta_final.resolve()),
                            }
                        )
                else:
                    filas_excel.append(
                        {
                            "fecha": fecha_emision,
                            "empresa": datos.get("empresa", ""),
                            "cantidad": "",
                            "descripcion": "",
                            "moneda": moneda,
                            "p_unit": "",
                            "total": "",
                            "url": str(ruta_final.resolve()),
                        }
                    )

            ruta_excel = base_facturas / "EXTRACCION_FACTURAS.xlsx"
            guardar_excel_por_anio_y_analisis(ruta_excel, filas_excel)
            log(f"Excel generado: {ruta_excel}")

            log("Descargas y extracción completadas.")
            log("Ahora el navegador NO se cerrará automáticamente.")

            mantener_navegador_abierto(browser, log_callback=log)

        except Exception as e:
            log(f"Error: {e}")
            raise


def _ir_a_consultar_factura(page, log):
    # En algunas vistas, volver a tocar el menú fuerza a mostrar el formulario de fechas nuevamente.
    ok = click_primer_visible(
        page,
        [
            ("locator", "span.spanNivelDescripcion:has-text('Consultar Factura y Nota')"),
            ("text", "Consultar Factura y Nota"),
        ],
        log=log,
        espera_ms=2500,
    )
    if not ok:
        raise Exception("No se pudo volver a 'Consultar Factura y Nota' para cambiar el mes.")


def ejecutar_rpa_por_meses(
    ruc,
    usuario,
    contrasena,
    desde_mes,
    desde_anio,
    hasta_mes,
    hasta_anio,
    log_callback=None,
):
    """
    Ejecuta el flujo mes a mes completo (01..último día), desde Mes/Año inicio hasta Mes/Año fin.
    Descarga PDFs y luego genera Excel por año + análisis.
    """

    def log(msg):
        if log_callback:
            log_callback(msg)
        else:
            print(msg)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False, slow_mo=250)

        base_facturas = Path(__file__).resolve().parent / "FACTURAS"
        base_facturas.mkdir(parents=True, exist_ok=True)

        context = browser.new_context(
            viewport={"width": 1366, "height": 900},
            locale="es-PE",
            accept_downloads=True,
        )
        page = context.new_page()

        filas_excel = []
        tmp_dir = base_facturas / "_tmp"
        tmp_dir.mkdir(parents=True, exist_ok=True)

        try:
            # Login + navegación hasta la pantalla de consulta (una sola vez)
            login_sunat(page, ruc, usuario, contrasena, log)
            log("Esperando menú principal...")
            page.wait_for_load_state("networkidle")
            page.wait_for_timeout(3000)

            log("Ingresando a Empresas...")
            if not click_primer_visible(page, [("locator", "h4:has-text('Empresas')"), ("text", "Empresas")], log=log):
                raise Exception("No se encontró la opción 'Empresas'.")

            log("Ingresando a Comprobantes de pago...")
            if not click_primer_visible(
                page,
                [
                    ("locator", "span.spanNivelDescripcion:has-text('Comprobantes de pago')"),
                    ("text", "Comprobantes de pago"),
                ],
                log=log,
            ):
                raise Exception("No se encontró la opción 'Comprobantes de pago'.")

            log("Ingresando a SEE - SOL...")
            if not click_primer_visible(
                page,
                [
                    ("locator", "span.spanNivelDescripcion:has-text('SEE - SOL')"),
                    ("text", "SEE - SOL"),
                ],
                log=log,
            ):
                raise Exception("No se encontró la opción 'SEE - SOL'.")

            log("Ingresando a Factura Electrónica...")
            if not click_primer_visible(
                page,
                [
                    ("locator", "span.spanNivelDescripcion:has-text('Factura Electrónica')"),
                    ("text", "Factura Electrónica"),
                    ("locator", "li#nivel3_11_5_3"),
                ],
                log=log,
            ):
                raise Exception("No se encontró la opción 'Factura Electrónica'.")

            log("Ingresando a Consultar Factura y Nota...")
            _ir_a_consultar_factura(page, log)

            # Si abre nueva pestaña
            if len(context.pages) > 1:
                page = context.pages[-1]
                page.wait_for_load_state("networkidle")
                page.wait_for_timeout(2000)

            # Procesar mes a mes
            for (fecha_inicio, fecha_fin, anio, mes_nombre, _mes_num) in _rango_mensual(
                desde_mes, desde_anio, hasta_mes, hasta_anio
            ):
                log(f"Procesando período: {fecha_inicio} - {fecha_fin} ({mes_nombre} {anio})")

                # Cerrar pestañas/popups sobrantes de la iteración anterior
                while len(context.pages) > 1:
                    try:
                        extra = context.pages[-1]
                        if extra != page:
                            extra.close()
                        else:
                            context.pages[0].close()
                    except Exception:
                        break

                # Volver a la opción para refrescar formulario
                try:
                    _ir_a_consultar_factura(page, log)
                except Exception:
                    # si ya estamos en la pantalla, ignorar
                    pass

                # Esperar a que aparezca el formulario de fechas (con reintentos)
                log("Buscando formulario de fechas...")
                target = None
                for _intento in range(6):
                    target = obtener_target_con_fechas(page)
                    if target is not None:
                        break
                    page.wait_for_timeout(1500)
                if target is None:
                    raise Exception("No se localizaron los campos de fecha.")

                # Limpiar los campos antes de rellenar
                log("Rellenando fechas...")
                campo_desde = target.locator('[id="criterio.fec_desde"]').first
                campo_hasta = target.locator('[id="criterio.fec_hasta"]').first
                campo_desde.click()
                campo_desde.fill("")
                campo_desde.fill(fecha_inicio)
                target.wait_for_timeout(400)
                campo_hasta.click()
                campo_hasta.fill("")
                campo_hasta.fill(fecha_fin)
                target.wait_for_timeout(700)

                log("Haciendo clic en Aceptar...")
                aceptado = click_primer_visible(
                    target,
                    [
                        ("role_button", "Aceptar"),
                        ("locator", "button:has-text('Aceptar')"),
                        ("locator", ".dijitButton:has(.nextIcon)"),
                        ("locator", "button:has(span.nextIcon)"),
                        ("text", "Aceptar"),
                    ],
                    log=log,
                    espera_ms=2500,
                )
                if not aceptado and target != page:
                    aceptado = click_primer_visible(
                        page,
                        [
                            ("role_button", "Aceptar"),
                            ("locator", "button:has-text('Aceptar')"),
                            ("locator", ".dijitButton:has(.nextIcon)"),
                            ("locator", "button:has(span.nextIcon)"),
                            ("text", "Aceptar"),
                        ],
                        log=log,
                        espera_ms=2500,
                    )
                if not aceptado:
                    raise Exception("No se encontró el botón 'Aceptar'.")

                # Esperar resultados usando la función robusta (polls para links PDF)
                log("Esperando resultados de la consulta...")
                try:
                    contenedor_res = esperar_resultados_consulta(page, log=log, timeout_seg=45)
                except Exception:
                    # Si no detecta resultados, puede ser que no haya facturas en el período
                    log("No se detectaron resultados para este período. Continuando al siguiente mes.")
                    continue

                time.sleep(1.5)

                # Buscar enlaces de descarga
                contenedor = contenedor_res if contenedor_res is not None else (target if target is not None else page)
                links = obtener_links_descargar_pdf(contenedor)
                if links is None or links.count() == 0:
                    # Intentar también en el target/frame
                    if contenedor != target and target is not None:
                        links = obtener_links_descargar_pdf(target)
                    if links is None or links.count() == 0:
                        links = obtener_links_descargar_pdf(page)
                if links is None or links.count() == 0:
                    log("No se encontraron 'Descargar PDF' en este período. Continuando al siguiente mes.")
                    continue

                total_links = links.count()
                log(f"Se encontraron {total_links} PDF(s) para {mes_nombre} {anio}.")

                for i in range(total_links):
                    ruta_tmp = descargar_pdf_desde_link(page, contenedor, i, tmp_dir, log)
                    datos = extraer_datos_factura_pdf(ruta_tmp)

                    empresa = _sanear_nombre_archivo(datos.get("empresa", "SIN_EMPRESA"))
                    serie = datos.get("serie", "")
                    numero = datos.get("numero", "")
                    fecha_emision = datos.get("fecha_emision", "")
                    moneda = datos.get("moneda", "")

                    carpeta_dest = base_facturas / anio / mes_nombre / empresa
                    carpeta_dest.mkdir(parents=True, exist_ok=True)

                    descripcion_primera = ""
                    items = datos.get("items") or []
                    if items:
                        descripcion_primera = items[0].get("descripcion", "")

                    desc_safe = _sanear_nombre_archivo(descripcion_primera).replace(" ", "_")
                    if serie and numero and desc_safe:
                        nombre_final = f"{serie}_{numero}_{desc_safe}.pdf"
                    elif serie and numero:
                        nombre_final = f"{serie}_{numero}.pdf"
                    else:
                        nombre_final = _sanear_nombre_archivo(ruta_tmp.stem) + ".pdf"

                    ruta_final = carpeta_dest / nombre_final
                    if ruta_final.exists():
                        base = ruta_final.stem
                        ext = ruta_final.suffix
                        k = 2
                        while True:
                            cand = carpeta_dest / f"{base}_{k}{ext}"
                            if not cand.exists():
                                ruta_final = cand
                                break
                            k += 1

                    ruta_tmp.replace(ruta_final)

                    if items:
                        for it in items:
                            filas_excel.append(
                                {
                                    "fecha": fecha_emision,
                                    "empresa": datos.get("empresa", ""),
                                    "cantidad": it.get("cantidad", ""),
                                    "descripcion": it.get("descripcion", ""),
                                    "moneda": moneda,
                                    "p_unit": it.get("p_unit", ""),
                                    "total": it.get("total", ""),
                                    "url": str(ruta_final.resolve()),
                                }
                            )
                    else:
                        filas_excel.append(
                            {
                                "fecha": fecha_emision,
                                "empresa": datos.get("empresa", ""),
                                "cantidad": "",
                                "descripcion": "",
                                "moneda": moneda,
                                "p_unit": "",
                                "total": "",
                                "url": str(ruta_final.resolve()),
                            }
                        )

            ruta_excel = base_facturas / "EXTRACCION_FACTURAS.xlsx"
            guardar_excel_por_anio_y_analisis(ruta_excel, filas_excel)
            log(f"Excel generado: {ruta_excel}")

            # Abrir carpeta FACTURAS y el Excel, independientemente de lo que elija luego.
            try:
                os.startfile(str(base_facturas))
            except Exception:
                log(f"No se pudo abrir la carpeta: {base_facturas}")

            try:
                os.startfile(str(ruta_excel))
            except Exception:
                log(f"No se pudo abrir el archivo de Excel: {ruta_excel}")

            # Preguntar si se desea cerrar el navegador
            cerrar = True
            try:
                cerrar = messagebox.askyesno(
                    "Proceso finalizado",
                    "Se terminó el proceso.\n\n¿Desea cerrar el navegador?",
                )
            except Exception:
                # Si hay algún problema con la UI, por seguridad se cierra.
                cerrar = True

            if cerrar:
                log("Cerrando navegador a solicitud del usuario.")
                try:
                    browser.close()
                except Exception:
                    pass
            else:
                log("Deja el navegador abierto. Ciérralo manualmente cuando desees.")
                mantener_navegador_abierto(browser, log_callback=log)

        except Exception as e:
            log(f"Error: {e}")
            raise


class AppSunat(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("SUNAT - Consulta de Facturas")
        self.geometry("560x500")
        self.minsize(520, 460)
        self.resizable(True, True)

        self._crear_estilos()
        self._crear_widgets()
        self._cargar_credenciales()

    def _crear_estilos(self):
        self.configure(bg="#f5f5f5")

        default_font = tkfont.nametofont("TkDefaultFont")
        default_font.configure(family="Segoe UI", size=10)

        text_font = tkfont.nametofont("TkTextFont")
        text_font.configure(family="Segoe UI", size=10)

        menu_font = tkfont.nametofont("TkMenuFont")
        menu_font.configure(family="Segoe UI", size=10)

        heading_font = tkfont.nametofont("TkHeadingFont")
        heading_font.configure(family="Segoe UI", size=10)

        fixed_font = tkfont.nametofont("TkFixedFont")
        fixed_font.configure(family="Consolas", size=9)

        self.font_titulo = tkfont.Font(family="Segoe UI", size=12, weight="bold")
        self.font_log = tkfont.Font(family="Consolas", size=9)

        style = ttk.Style(self)
        try:
            style.theme_use("vista")
        except Exception:
            pass

        style.configure("TLabel", background="#f5f5f5")
        style.configure("TFrame", background="#f5f5f5")
        style.configure("TLabelframe", background="#f5f5f5")
        style.configure("TLabelframe.Label", font=self.font_titulo)
        style.configure("TButton", padding=6)

    def _crear_widgets(self):
        main = ttk.Frame(self, padding=20)
        main.pack(fill=tk.BOTH, expand=True)

        lbl_creds = ttk.Label(main, text="Credenciales SUNAT SOL", font=self.font_titulo)
        lbl_creds.pack(anchor=tk.W, pady=(0, 12))

        frm_creds = ttk.LabelFrame(main, text="Datos de acceso", padding=10)
        frm_creds.pack(fill=tk.X, pady=(0, 12))

        ttk.Label(frm_creds, text="RUC:").grid(row=0, column=0, sticky=tk.W, pady=4, padx=(0, 8))
        self.var_ruc = tk.StringVar()
        self.ent_ruc = ttk.Entry(frm_creds, textvariable=self.var_ruc, width=20)
        self.ent_ruc.grid(row=0, column=1, sticky=tk.EW, pady=4)
        frm_creds.columnconfigure(1, weight=1)

        ttk.Label(frm_creds, text="Usuario:").grid(row=1, column=0, sticky=tk.W, pady=4, padx=(0, 8))
        self.var_usuario = tk.StringVar()
        self.ent_usuario = ttk.Entry(frm_creds, textvariable=self.var_usuario, width=20)
        self.ent_usuario.grid(row=1, column=1, sticky=tk.EW, pady=4)

        ttk.Label(frm_creds, text="Contraseña:").grid(row=2, column=0, sticky=tk.W, pady=4, padx=(0, 8))
        self.var_contrasena = tk.StringVar()
        self.ent_contrasena = ttk.Entry(frm_creds, textvariable=self.var_contrasena, show="*", width=20)
        self.ent_contrasena.grid(row=2, column=1, sticky=tk.EW, pady=4)

        lbl_fechas = ttk.Label(main, text="Rango de fechas para consulta", font=self.font_titulo)
        lbl_fechas.pack(anchor=tk.W, pady=(12, 8))

        # --- Radio buttons: Anual / Mensual ---
        frm_modo = ttk.Frame(main)
        frm_modo.pack(fill=tk.X, pady=(0, 6))

        self.var_modo_fecha = tk.StringVar(value="anual")
        rb_anual = ttk.Radiobutton(frm_modo, text="Anual", variable=self.var_modo_fecha, value="anual", command=self._toggle_modo_fecha)
        rb_anual.pack(side=tk.LEFT, padx=(0, 16))
        rb_mensual = ttk.Radiobutton(frm_modo, text="Mensual", variable=self.var_modo_fecha, value="mensual", command=self._toggle_modo_fecha)
        rb_mensual.pack(side=tk.LEFT)

        # --- Contenedor fijo para los frames de fecha (mantiene posición en el layout) ---
        self.frm_fechas_container = ttk.Frame(main)
        self.frm_fechas_container.pack(fill=tk.X, pady=(0, 12))

        # --- Frame ANUAL (Mes/Año combos) ---
        self.frm_anual = ttk.LabelFrame(self.frm_fechas_container, text="Rango por Año y Mes", padding=10)

        ttk.Label(self.frm_anual, text="Mes/Año Inicio:").grid(row=0, column=0, sticky=tk.W, pady=4, padx=(0, 8))
        self.var_mes_ini = tk.StringVar(value="ENERO")
        self.var_anio_ini = tk.StringVar(value=str(datetime.now().year))
        self.cmb_mes_ini = ttk.Combobox(self.frm_anual, textvariable=self.var_mes_ini, values=MESES_LISTA, state="readonly", width=14)
        self.cmb_mes_ini.grid(row=0, column=1, sticky=tk.W, pady=4)
        self.ent_anio_ini = ttk.Entry(self.frm_anual, textvariable=self.var_anio_ini, width=8)
        self.ent_anio_ini.grid(row=0, column=2, sticky=tk.W, pady=4, padx=(8, 0))

        ttk.Label(self.frm_anual, text="Mes/Año Fin:").grid(row=1, column=0, sticky=tk.W, pady=4, padx=(0, 8))
        self.var_mes_fin = tk.StringVar(value="ENERO")
        self.var_anio_fin = tk.StringVar(value=str(datetime.now().year))
        self.cmb_mes_fin = ttk.Combobox(self.frm_anual, textvariable=self.var_mes_fin, values=MESES_LISTA, state="readonly", width=14)
        self.cmb_mes_fin.grid(row=1, column=1, sticky=tk.W, pady=4)
        self.ent_anio_fin = ttk.Entry(self.frm_anual, textvariable=self.var_anio_fin, width=8)
        self.ent_anio_fin.grid(row=1, column=2, sticky=tk.W, pady=4, padx=(8, 0))

        # --- Frame MENSUAL (Calendarios) ---
        self.frm_mensual = ttk.LabelFrame(self.frm_fechas_container, text="Rango por Calendario (Mes Inicio - Mes Fin)", padding=10)

        ttk.Label(self.frm_mensual, text="Mes Inicio:").grid(row=0, column=0, sticky=tk.W, pady=4, padx=(0, 8))
        self.cal_mes_ini = DateEntry(self.frm_mensual, width=14, locale="es_PE", date_pattern="dd/mm/yyyy")
        self.cal_mes_ini.grid(row=0, column=1, sticky=tk.W, pady=4)

        ttk.Label(self.frm_mensual, text="Mes Fin:").grid(row=1, column=0, sticky=tk.W, pady=4, padx=(0, 8))
        self.cal_mes_fin = DateEntry(self.frm_mensual, width=14, locale="es_PE", date_pattern="dd/mm/yyyy")
        self.cal_mes_fin.grid(row=1, column=1, sticky=tk.W, pady=4)

        # Mostrar el frame correcto al inicio
        self._toggle_modo_fecha()

        frm_log = ttk.LabelFrame(main, text="Estado", padding=6)
        frm_log.pack(fill=tk.BOTH, expand=True, pady=(0, 12))

        self.txt_log = tk.Text(
            frm_log,
            height=8,
            wrap=tk.WORD,
            state=tk.DISABLED,
            font=self.font_log
        )
        self.txt_log.pack(fill=tk.BOTH, expand=True)

        self.btn_proceder = ttk.Button(
            main,
            text="Proceder",
            command=self._on_proceder,
        )
        self.btn_proceder.pack(pady=8, ipadx=20, ipady=6)

    def _cargar_credenciales(self):
        creds = cargar_credenciales_desde_archivo()
        if creds:
            self.var_ruc.set(creds.get("ruc", ""))
            self.var_usuario.set(creds.get("usuario", ""))
            self.var_contrasena.set(creds.get("contrasena", ""))

    def _append_log_ui(self, msg):
        self.txt_log.configure(state=tk.NORMAL)
        self.txt_log.insert(tk.END, msg + "\n")
        self.txt_log.see(tk.END)
        self.txt_log.configure(state=tk.DISABLED)

    def _log(self, msg):
        self.after(0, self._append_log_ui, msg)

    def _habilitar_boton(self):
        self.btn_proceder.configure(state=tk.NORMAL)

    def _deshabilitar_boton(self):
        self.btn_proceder.configure(state=tk.DISABLED)

    def _toggle_modo_fecha(self):
        """Muestra u oculta los frames de fecha según el modo seleccionado."""
        modo = self.var_modo_fecha.get()
        if modo == "anual":
            self.frm_mensual.pack_forget()
            self.frm_anual.pack(in_=self.frm_fechas_container, fill=tk.X)
        else:
            self.frm_anual.pack_forget()
            self.frm_mensual.pack(in_=self.frm_fechas_container, fill=tk.X)

    def _on_proceder(self):
        ruc = self.var_ruc.get().strip()
        usuario = self.var_usuario.get().strip()
        contrasena = self.var_contrasena.get()

        if not ruc:
            messagebox.showerror("Error", "Ingresa el RUC.")
            return
        if not usuario:
            messagebox.showerror("Error", "Ingresa el usuario.")
            return
        if not contrasena:
            messagebox.showerror("Error", "Ingresa la contraseña.")
            return

        modo = self.var_modo_fecha.get()

        if modo == "anual":
            mes_ini = (self.var_mes_ini.get() or "").strip().upper()
            mes_fin = (self.var_mes_fin.get() or "").strip().upper()
            anio_ini = (self.var_anio_ini.get() or "").strip()
            anio_fin = (self.var_anio_fin.get() or "").strip()

            if mes_ini not in MESES_IDX or mes_fin not in MESES_IDX:
                messagebox.showerror("Error", "Selecciona Mes/Año de inicio y fin.")
                return
            if not (anio_ini.isdigit() and len(anio_ini) == 4 and anio_fin.isdigit() and len(anio_fin) == 4):
                messagebox.showerror("Error", "Año inválido. Usa formato AAAA (ej: 2025).")
                return

            desde_mes = MESES_IDX[mes_ini]
            hasta_mes = MESES_IDX[mes_fin]
            desde_anio = int(anio_ini)
            hasta_anio = int(anio_fin)
        else:
            # Modo mensual: leer fechas del calendario
            fecha_ini = self.cal_mes_ini.get_date()
            fecha_fin = self.cal_mes_fin.get_date()
            desde_mes = fecha_ini.month
            desde_anio = fecha_ini.year
            hasta_mes = fecha_fin.month
            hasta_anio = fecha_fin.year

        guardar_credenciales(ruc, usuario, contrasena)

        self._deshabilitar_boton()
        self.txt_log.configure(state=tk.NORMAL)
        self.txt_log.delete("1.0", tk.END)
        self.txt_log.configure(state=tk.DISABLED)

        def run_rpa():
            try:
                self._log("Iniciando flujo...")
                self._log("Nota: el botón se habilitará nuevamente cuando cierres el navegador.")
                ejecutar_rpa_por_meses(
                    ruc=ruc,
                    usuario=usuario,
                    contrasena=contrasena,
                    desde_mes=desde_mes,
                    desde_anio=desde_anio,
                    hasta_mes=hasta_mes,
                    hasta_anio=hasta_anio,
                    log_callback=self._log,
                )
            except Exception as e:
                self.after(0, lambda: messagebox.showerror("Error", str(e)))
            finally:
                self.after(0, self._habilitar_boton)

        threading.Thread(target=run_rpa, daemon=True).start()


def main():
    app = AppSunat()
    app.mainloop()


if __name__ == "__main__":
    main()
